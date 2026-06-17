#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
电力市场用电侧中长期超缺额偏差收益回收计算脚本

功能：根据用户提供的"用电侧中长期签约缺额（超额）偏差收益回收.xlsx"数据，
     自动计算全月各时段的偏差收益回收金额，并输出逐时段明细和汇总表到 Excel。

业务逻辑：
  - 自动检测月份天数（28~31天）和每日时段数（24/48/96等）
  - 自动定位各数据块标题行（不依赖固定行号）
  - 对每个时段，判断合约电量相对于实际用电量属于"缺额偏差"、"超额偏差"还是"没有偏差"
  - 按规定公式计算偏差收益回收金额

v2.0 变更：
  - 新增动态数据块定位，不再依赖固定行间距（34/35行）
  - 适配集中交易电量/价格行数不确定的场景（如多次月内交易叠加导致464行+）
  - 自动检测实时市场出清块位置（可能在集中交易块之前或之后）

作者：智测助手 🦞
"""

import sys
import os
import openpyxl
import numpy as np
import datetime
import re
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# 数据读取函数
# =============================================================================

def read_matrix_from_values(all_values, sr, cs, rows, periods):
    """
    从预加载的 values 列表中读取二维数据矩阵（高性能版本）。

    参数：
        all_values — list(ws.values)，所有行数据（0-based 索引）
        sr       — 起始行号（1-based，即Excel行号，通常为标题行）
        cs       — 起始列号（1-based，即Excel列号）
        rows     — 数据行数
        periods  — 每日时段数（如24表示24个时段）

    返回：
        numpy 二维数组，shape = (rows, periods)

    说明：
        读取 Excel 中 [sr+1, sr+rows] × [cs, cs+periods-1] 的区域。
        单元格值为 None、datetime.time 或非数字时视为 0.0。
    """
    data = []
    # sr 是 1-based 标题行，sr+1 是数据第一行（1-based）
    # 转换为 0-based: 数据第一行 = sr（因为 sr+1-1 = sr）
    start_row = sr  # 0-based
    for d in range(rows):
        row_data = all_values[start_row + d]
        row = []
        for h in range(periods):
            if cs + h - 1 < len(row_data):
                v = row_data[cs + h - 1]  # cs 是 1-based，转 0-based
            else:
                v = None
            if v is None:
                row.append(0.0)
            elif isinstance(v, datetime.time):
                row.append(0.0)
            else:
                try:
                    row.append(float(v))
                except (ValueError, TypeError):
                    row.append(0.0)
        data.append(row)
    return np.array(data)


def _is_valid_day_label(v):
    """检查单元格值是否为合法的天数标签（1-31 的整数或等价浮点数）。"""
    if v is None:
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return 1 <= v <= 31
    if isinstance(v, float):
        return v == int(v) and 1 <= int(v) <= 31
    if isinstance(v, str):
        try:
            n = int(v)
            return 1 <= n <= 31
        except (ValueError, TypeError):
            return False
    return False


def _is_row_all_numbers(ws, row):
    """判断一行是否整行都是数字（用于识别"列序号行"等异常行）。"""
    max_col = ws.max_column or 50
    non_empty = 0
    for c in range(1, min(max_col, 60)):
        v = ws.cell(row=row, column=c).value
        if v is not None:
            non_empty += 1
            is_num = False
            if isinstance(v, bool):
                is_num = False
            elif isinstance(v, (int, float)):
                is_num = True
            elif isinstance(v, str):
                try:
                    float(v)
                    is_num = True
                except (ValueError, TypeError):
                    pass
            if not is_num:
                return False
    return non_empty >= 15


def auto_detect_days(ws):
    """
    自动检测 Excel 数据中的月份天数。
    从第3行起扫描A列，寻找1-31的天数标签。
    注意：数据行A列的1-31是合法天数标签，不触发"整行数字"排除。
    "整行数字"检查仅用于排除A列非1-31的异常行（如列序号行）。
    """
    days = 0
    for d in range(40):
        row = 3 + d
        v = ws.cell(row=row, column=1).value
        if v is None:
            break
        if not _is_valid_day_label(v):
            break
        # A列已经是合法天数标签（1-31），无需再做"整行数字"检查
        # （数据行B-Y列全是电量数字，会误触发）
        days += 1
    return days


def auto_detect_periods(ws):
    """
    自动检测每日时段数（如24即表示一天分为24个时段）。
    在前10行中寻找第一个 datetime.time 类型的时间值，向右统计。
    注意：时间值可能从列2（B列）或列3（C列）开始，都支持。
    """
    hr = None
    start_col = None
    for r in range(1, 10):
        for c in range(2, 10):  # 从列2开始扫（B列可能是00:00）
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, datetime.time):
                hr = r
                start_col = c
                break
        if hr:
            break

    if not hr:
        return 0

    p = 0
    for c in range(start_col, 300):
        v = ws.cell(row=hr, column=c).value
        if v is not None and isinstance(v, datetime.time):
            p += 1
        elif p > 0:
            break
    return p


# =============================================================================
# v2.0 新增：动态数据块定位
# =============================================================================

def find_block_start(ws, keyword, max_search=500):
    """
    在 Excel 工作表中扫描包含指定关键词的标题行。

    扫描范围：行1 ~ max_search，检查 A、AA、AB 列。

    参数：
        ws          — Excel 工作表对象
        keyword     — 标题关键词（如"集中交易电量"）
        max_search  — 最大扫描行数

    返回：
        标题行号（1-based），找不到返回 None
    """
    search_cols = [1, 27, 28]  # A, AA, AB
    for r in range(1, min(max_search, ws.max_row) + 1):
        for c in search_cols:
            if c > ws.max_column:
                continue
            v = ws.cell(row=r, column=c).value
            if v is not None and keyword in str(v):
                return r
    return None


def count_block_rows(ws, title_row, max_periods=2000):
    """
    从数据块标题行向下计算数据行数。

    有效数据行判定（A列或AA列满足其一即可）：
      - 1~31 的整数/浮点数（天数标签）
      - ISO 日期字符串（如 "2026-03-01"）

    连续出现 ≥5 行无效数据时终止。

    参数：
        ws          — Excel 工作表对象
        title_row   — 数据块标题行号（1-based）
        max_periods — 最大扫描行数（防止溢出）

    返回：
        数据行数（整数）
    """
    def _valid_day(v):
        if v is None:
            return False
        if _is_valid_day_label(v):
            return True
        if isinstance(v, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', v.strip()):
            return True
        return False

    count = 0
    invalid_streak = 0
    max_invalid = 5

    for offset in range(1, max_periods + 1):
        row = title_row + offset
        a_val = ws.cell(row=row, column=1).value
        aa_val = ws.cell(row=row, column=27).value

        if _valid_day(a_val) or _valid_day(aa_val):
            count += 1
            invalid_streak = 0
        else:
            invalid_streak += 1
            if invalid_streak >= max_invalid:
                break

    return count if count > 0 else 1


def is_single_row_block(ws, title_row, first_col, periods):
    """
    判断一个数据块是否只有1行数据（福建版均价块的特征）。

    判定规则：标题行下方第1行是时间值行（包含 datetime.time），
    第2行起连续5行都不是有效天数标签 → 判定为单行数据块。

    返回：(is_single, data_row) — is_single为True时，data_row为数据行号（1-based）
    """
    def _valid_day(v):
        if v is None:
            return False
        if _is_valid_day_label(v):
            return True
        if isinstance(v, str) and re.match(r'^\d{4}-\d{2}-\d{2}$', v.strip()):
            return True
        return False

    # 检查第2行是否为时间行（标题行下方第1行是列头时间行）
    row_below_title = title_row + 1
    has_time = False
    for c in range(first_col, first_col + periods + 2):
        if c > 60:
            break
        v = ws.cell(row=row_below_title, column=c).value
        if v is not None and isinstance(v, datetime.time):
            has_time = True
            break

    if not has_time:
        return False, None

    # 检查第3行起是否不是天数标签（连续3行即可判定）
    for offset in range(2, 5):
        row = title_row + offset
        a_val = ws.cell(row=row, column=1).value
        if _valid_day(a_val):
            return False, None  # 找到了天数标签，说明不是单行块

    return True, row_below_title


def find_first_data_col(ws, max_rows=10):
    """
    动态检测电量数据的起始列。

    在前max_rows行中，找到第一个 datetime.time 值所在的列，
    该列即为00:00对应的数据起始列。

    兜底：如果找不到时间值，返回3（兼容旧版格式：A列无日期时数据从C列起）。
    """
    for r in range(1, max_rows + 1):
        for c in range(1, 10):
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, datetime.time) and v.hour == 0:
                return c
    return 3  # 兜底：旧版格式从C列起


def find_price_col(ws, hr, periods, first_col):
    """
    定位"价格列"的起始列号。

    在列头行（hr）中寻找"类型"字符串，找到后在其后第一个时间值所在列，
    即为价格列的起始位置。

    兜底：如果找不到"类型"关键字，按 first_col + periods + 1 计算
    （first_col是电量起始列，右侧隔1列为价格起始列，适配AA列="日期"布局）。
    """
    found_type = False
    for c in range(3, 250):
        if c > ws.max_column:
            break
        v = ws.cell(row=hr, column=c).value
        if v == '类型':
            found_type = True
        elif found_type and isinstance(v, datetime.time):
            return c
    # 兜底：first_col + periods + 2（如B列起电量→24列→AA列日期→AB列价格，中间隔1列"日期"）
    return first_col + periods + 2


# =============================================================================
# 加权均价与辅助计算函数
# =============================================================================

def weighted_avg_price(qty, price):
    """
    计算分时加权均价（核心公式）。
    公式：分时加权均价[t] = Σ(d)[qty[d][t] × price[d][t]] / Σ(d)[qty[d][t]]
    """
    return (qty * price).sum(axis=0) / qty.sum(axis=0)


# =============================================================================
# 数据模式检测
# =============================================================================

def detect_data_mode(ws, first_col, periods):
    """
    检测 Excel 数据形态，区分三种模式：

    1. 多行电量+价格（need_weighted=True）→ 需加权计算均价
       特征：找到「集中交易电量」+「实时市场出清电量」多行数据块

    2. 单行分时均价（need_weighted=False, found_avg_blocks）
       特征：标题行+1 是时间行，标题行+2 不是天数标签 → 单行均价直接读取

    3. 同行并排均价（need_weighted=False, side_by_side_blocks）
       特征：标题行+1 是时间行，标题行+2 是天数标签 → 价格在电量右侧列，逐日逐时段

    返回：
        (need_weighted, found_avg_blocks, side_by_side_blocks)
    """
    avg_price_keywords = [
        '批发侧合同结算均价',
        '现货实时市场加权平均价',
        '集中交易加权均价',
        '实时市场加权均价',
        '集中交易价格',
        '实时结算电价',
    ]

    found_avg_blocks = []
    side_by_side_blocks = []

    for kw in avg_price_keywords:
        title_row = find_block_start(ws, kw)
        if title_row is None:
            continue

        is_single, data_row = is_single_row_block(ws, title_row, first_col, periods)
        if is_single:
            found_avg_blocks.append((kw, title_row, data_row))
            continue

        # side-by-side 检测：标题行+1=时间行，标题行+2=天数标签
        title_col = None
        for c in range(1, min(ws.max_column, 60)):
            v = ws.cell(row=title_row, column=c).value
            if v is not None and kw in str(v):
                title_col = c
                break
        if title_col is None:
            continue

        has_time = False
        for c in range(title_col + 1, title_col + periods + 3):
            v = ws.cell(row=title_row + 1, column=c).value
            if v is not None and isinstance(v, datetime.time):
                has_time = True
                break
        if not has_time:
            continue

        day_v = ws.cell(row=title_row + 2, column=1).value
        if _is_valid_day_label(day_v):
            price_start_col = title_col + 1
            data_start = 3 if title_row <= 20 else title_row + 2
            side_by_side_blocks.append((kw, title_row, price_start_col, data_start))

    need_weighted = len(found_avg_blocks) < 2 and len(side_by_side_blocks) < 2
    return need_weighted, found_avg_blocks, side_by_side_blocks


# =============================================================================
# 加权均价与辅助计算函数
# =============================================================================

def _judge_and_calc(a, c, pd, lambda_val, alpha3, alpha4):
    """
    三段式偏差判断与回收金额计算（四川/福建通用）。

    参数：
        a — 实际用电量（四川为全月汇总值，福建为单日单时段值）
        c — 合约电量（四川为全月汇总值，福建为单日单时段值）
        pd — 价差（实时价 − 集中价/合同价）
    返回：
        (rec, cond) — 回收金额、触发条件
    """
    if c < a * alpha3:
        rec = lambda_val * max(0, -pd) * (a * alpha3 - c)
        cond = "缺额回收"
    elif c > a * alpha4:
        rec = lambda_val * max(0, pd) * (c - a * alpha4)
        cond = "超额回收"
    else:
        rec = 0.0
        cond = "没有偏差"
    return rec, cond


# =============================================================================
# 核心计算逻辑
# =============================================================================

def calculate(file_path, lambda_val=1.1, alpha3=0.85, alpha4=1.15):
    """
    读取 Excel 数据，执行偏差收益回收计算，返回所有时段的计算结果。

    实现步骤：
      1. 自动检测天数和时段数
      2. 动态定位各数据块标题行，计算实际数据行数
      3. 读取6个数据块：用户实际用电量、中长期合约电量、集中交易电量、
         集中交易价格、实时市场出清电量、实时结算电价
      4. 计算实时分时加权均价和集中交易分时加权均价
      5. 对每个时段进行三段式偏差判断（缺额/超额/未触发），
         按公式计算偏差收益回收金额

    参数：
        file_path  — Excel 文件路径
        lambda_val — 考核系数 λ（默认 1.1）
        alpha3     — 缺额偏差触发线 α₃（默认 0.85）
        alpha4     — 超额偏差触发线 α₄（默认 1.15）
    返回：(results, periods) — results为每时段的详细结果列表，periods为时段数
    """
    # ── 输入校验 ──
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"输入文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    if '输入数据' not in wb.sheetnames:
        raise ValueError(f"找不到 Sheet「输入数据」，可用 Sheet: {wb.sheetnames}")
    ws = wb['输入数据']

    # -------------------------------------------------------------------------
    # Step 1: 自动检测天数和时段数
    # -------------------------------------------------------------------------
    days = auto_detect_days(ws)
    periods = auto_detect_periods(ws)
    if days == 0:
        raise ValueError("无法检测天数")
    if periods == 0:
        raise ValueError("无法检测时段数")
    print(f"[INFO] 自动检测: {days}天 x {periods}时段")
    print(f"[INFO] 参数：λ={lambda_val}, α₃={alpha3}, α₄={alpha4}")

    # -------------------------------------------------------------------------
    # Step 2: 动态检测起始列 + 定位价格列
    # -------------------------------------------------------------------------
    hr = None
    for r in range(1, 10):
        for c in range(2, 10):  # 从列2开始扫（B列可能是00:00）
            v = ws.cell(row=r, column=c).value
            if v is not None and isinstance(v, datetime.time):
                hr = r
                break
        if hr:
            break

    first_col = find_first_data_col(ws)  # 动态检测电量起始列（2=B列或3=C列）
    price_col = find_price_col(ws, hr, periods, first_col)
    print(f"[INFO] 电量起始列={first_col}, 价格起始列={price_col}")

    # -------------------------------------------------------------------------
    # Step 2.5: 动态定位各数据块（v2.0）
    # -------------------------------------------------------------------------
    print("[INFO] 动态定位数据块...")

    # 用户实际用电量：动态扫描（find_block_start 返回标题行，read_matrix 需要列头行=标题行+1）
    user_title_block = find_block_start(ws, '用户实际用电量')
    if user_title_block is None:
        user_title_block = 1  # 兜底：固定行1
    user_title = user_title_block + 1  # 列头行

    # 中长期合约电量：动态扫描（find_block_start 返回的是标题行，read_matrix 需要列头行=标题行+1）
    contract_title_block = find_block_start(ws, '中长期合约')
    if contract_title_block is None:
        raise ValueError("未找到「中长期合约电量」标题行")
    contract_title = contract_title_block + 1  # 列头行（如行36=标题 → 行37=列头）

    # ---------------------------------------------------------------------
    # 版本检测：按数据形态区分
    # ---------------------------------------------------------------------
    need_weighted, found_avg_blocks, side_by_side_blocks = detect_data_mode(
        ws, first_col, periods)
    if not need_weighted:
        if found_avg_blocks:
            print(f"[INFO] 检测到单行分时均价数据块（{len(found_avg_blocks)}个），跳过加权均价计算")
            for kw, tr, dr in found_avg_blocks:
                print(f"[INFO] 均价块: {kw}, 标题行{tr}, 数据行{dr}")
        if side_by_side_blocks:
            print(f"[INFO] 检测到同行并排均价数据块（{len(side_by_side_blocks)}个），跳过加权均价计算")
            for kw, tr, pc, ds in side_by_side_blocks:
                print(f"[INFO] 均价块: {kw}, 标题行{tr}, 价格列{pc}, 数据起始行{ds}")

    # 集中交易电量：动态扫描
    spot_title = find_block_start(ws, '集中交易电量')

    # 实时市场出清电量：动态扫描
    clearing_title = find_block_start(ws, '实时市场出清电量')

    if need_weighted:
        if spot_title is None:
            raise ValueError("未找到「集中交易电量」标题行，也未检测到单行均价数据块")
        spot_rows = count_block_rows(ws, spot_title)
        print(f"[INFO] 集中交易块: 标题行{spot_title}, 数据{spot_rows}行")
        if clearing_title is None:
            raise ValueError("未找到「实时市场出清电量」标题行")
        print(f"[INFO] 实时市场块: 标题行{clearing_title}")

    # -------------------------------------------------------------------------
    # Step 3: 读取6个数据块（预加载整表，批量读取）
    # -------------------------------------------------------------------------
    all_values = list(ws.values)  # 一次性加载整表到内存

    user_q = read_matrix_from_values(all_values, user_title, first_col, days, periods)            # 用户实际用电量
    contract_q = read_matrix_from_values(all_values, contract_title, first_col, days, periods)     # 中长期合约电量

    # -------------------------------------------------------------------------
    # Step 3.5: 根据数据形态读取均价数据
    # -------------------------------------------------------------------------
    if not need_weighted:
        def read_avg_row(data_row, fc, p):
            """从单行数据中读取分时段均价数组。"""
            vals = []
            for h in range(p):
                c_idx = fc + h - 1  # 转0-based列索引
                if data_row - 1 < len(all_values) and c_idx < len(all_values[data_row - 1]):
                    v = all_values[data_row - 1][c_idx]
                else:
                    v = None
                if v is None:
                    vals.append(0.0)
                else:
                    try:
                        vals.append(float(v))
                    except (ValueError, TypeError):
                        vals.append(0.0)
            return np.array(vals)

        def read_price_matrix(data_row_start, pc, days, p):
            """从side-by-side格式中读取逐日逐时段价格矩阵。
            返回 (days, periods) 矩阵，每个(d, t)都有对应的价格。"""
            matrix = np.zeros((days, p))
            for d in range(days):
                r_idx = data_row_start + d - 1  # 1-based → 0-based
                if r_idx < len(all_values):
                    row = all_values[r_idx]
                    for h in range(p):
                        c_idx = pc + h - 1  # 1-based → 0-based
                        if c_idx < len(row) and row[c_idx] is not None:
                            try:
                                matrix[d][h] = float(row[c_idx])
                            except (ValueError, TypeError):
                                pass
            return matrix  # (days, periods) 完整矩阵

        conc_avg = None
        real_avg = None
        conc_price_matrix = None  # (days, periods) 福建模式用
        real_price_matrix = None  # (days, periods) 福建模式用

        # 先尝试从单行均价块读取（四川/其他省份的单行模式）
        if found_avg_blocks:
            for kw, tr, dr in found_avg_blocks:
                if kw in ('批发侧合同结算均价', '集中交易加权均价', '集中交易价格'):
                    conc_avg = read_avg_row(dr, first_col, periods)
                elif kw in ('现货实时市场加权平均价', '实时市场加权均价', '实时结算电价'):
                    real_avg = read_avg_row(dr, first_col, periods)

        # 如果单行块不够，尝试从side-by-side块读取（福建模式）
        if conc_avg is None or real_avg is None:
            for kw, tr, pc, ds in side_by_side_blocks:
                if kw in ('批发侧合同结算均价', '集中交易加权均价', '集中交易价格') and conc_avg is None:
                    conc_price_matrix = read_price_matrix(ds, pc, days, periods)
                    conc_avg = conc_price_matrix.mean(axis=0)  # 导出用均值
                elif kw in ('现货实时市场加权平均价', '实时市场加权均价', '实时结算电价') and real_avg is None:
                    real_price_matrix = read_price_matrix(ds, pc, days, periods)
                    real_avg = real_price_matrix.mean(axis=0)  # 导出用均值

        if conc_avg is None or real_avg is None:
            raise ValueError(f"无法读取均价数据: found_avg_blocks={len(found_avg_blocks)}, side_by_side_blocks={len(side_by_side_blocks)}")

        if conc_price_matrix is not None or real_price_matrix is not None:
            print(f"[INFO] 逐日逐时段价格矩阵读取完成: 集中({days}天×{periods}时段), 实时({days}天×{periods}时段)")
        else:
            print(f"[INFO] 均价数据读取完成（直接读取）: 集中均价({periods}时段), 实时均价({periods}时段)")
    else:
        # 多行电量+价格块 → 读取后计算加权均价
        spot_q = read_matrix_from_values(all_values, spot_title + 1, first_col, spot_rows, periods)        # 集中交易电量
        conc_price = read_matrix_from_values(all_values, spot_title + 1, price_col, spot_rows, periods)    # 集中交易价格
        clearing_q = read_matrix_from_values(all_values, clearing_title + 1, first_col, days, periods)     # 实时市场出清电量
        real_price = read_matrix_from_values(all_values, clearing_title + 1, price_col, days, periods)     # 实时结算电价
        print(f"[INFO] 电量/价格数据读取完成: 用户{user_q.shape}, 合约{contract_q.shape}, "
              f"集中交易{spot_q.shape}, 出清{clearing_q.shape}")

    # -------------------------------------------------------------------------
    # Step 4: 计算分时加权均价（仅需要加权计算时执行）
    # -------------------------------------------------------------------------
    if need_weighted:
        real_avg = weighted_avg_price(clearing_q, real_price)
        conc_avg = weighted_avg_price(spot_q, conc_price)
        print(f"[INFO] 分时加权均价计算完成")

    # -------------------------------------------------------------------------
    # Step 5: 三段式偏差判断与偏差收益回收计算
    # 四川（need_weighted=True）：全月分时汇总判定 — 按每个时段t汇总所有天d
    # 福建（need_weighted=False）：逐日逐时段判定 — 每个(d, t)独立判定
    # -------------------------------------------------------------------------
    results = []

    if need_weighted:
        # ── 四川模式：全月分时汇总判定 ──
        pu = user_q.sum(axis=0)     # 各时段实际用电汇总 (periods,)
        pc = contract_q.sum(axis=0) # 各时段合约电量汇总 (periods,)

        for t in range(periods):
            a = pu[t]
            c = pc[t]
            ratio = c / a if a > 0 else 0
            pd = real_avg[t] - conc_avg[t]

            rec, cond = _judge_and_calc(a, c, pd, lambda_val, alpha3, alpha4)

            results.append({
                "day": None,
                "t": t + 1,
                "actual": round(a, 2),
                "contract": round(c, 2),
                "ratio": round(ratio, 4),
                "realtime_avg": round(real_avg[t], 2),
                "concentrate_avg": round(conc_avg[t], 2),
                "price_diff": round(pd, 2),
                "condition": cond,
                "recovery": round(rec, 2)
            })
    else:
        # ── 福建模式：逐日逐时段独立判定 ──
        for d in range(days):
            for t in range(periods):
                a = user_q[d][t]
                c = contract_q[d][t]
                ratio = c / a if a > 0 else 0
                # 使用逐日逐时段的实际价格计算价差
                if real_price_matrix is not None and conc_price_matrix is not None:
                    pd = real_price_matrix[d][t] - conc_price_matrix[d][t]
                else:
                    pd = real_avg[t] - conc_avg[t]

                rec, cond = _judge_and_calc(a, c, pd, lambda_val, alpha3, alpha4)

                results.append({
                    "day": d + 1,
                    "t": t + 1,
                    "actual": round(a, 2),
                    "contract": round(c, 2),
                    "ratio": round(ratio, 4),
                    "realtime_avg": round(real_price_matrix[d][t] if real_price_matrix is not None else real_avg[t], 2),
                    "concentrate_avg": round(conc_price_matrix[d][t] if conc_price_matrix is not None else conc_avg[t], 2),
                    "price_diff": round(pd, 2),
                    "condition": cond,
                    "recovery": round(rec, 2)
                })

    return results, periods, need_weighted


# =============================================================================
# 控制台输出函数
# =============================================================================

def print_results(results):
    """在终端打印逐时段明细表和月度汇总。"""
    has_day = any(r.get("day") is not None for r in results)
    deficit_total = sum(r["recovery"] for r in results if r["condition"] == "缺额回收")
    excess_total = sum(r["recovery"] for r in results if r["condition"] == "超额回收")
    total = deficit_total + excess_total

    if has_day:
        # 福建模式：逐日逐时段
        print(f"\n日期  时段  触发类型  实际用电   合约电量     比率   回收额")
        print("-" * 78)
        for r in results:
            print(f"  {r['day']:>3}  {r['t']:>3}  {r['condition']:>6}  {r['actual']:>10.2f}  "
                  f"{r['contract']:>10.2f}  {r['ratio'] * 100:>7.2f}%  {r['recovery']:>14.2f}")
    else:
        # 四川模式：全月分时汇总
        print(f"\n时段  触发类型  实际用电   合约电量     比率   回收额")
        print("-" * 70)
        for r in results:
            print(f"  {r['t']:>3}  {r['condition']:>6}  {r['actual']:>10.2f}  "
                  f"{r['contract']:>10.2f}  {r['ratio'] * 100:>7.2f}%  {r['recovery']:>14.2f}")
    print("-" * 78 if has_day else "-" * 70)
    print(f"缺额回收合计: {deficit_total:>14.2f}")
    print(f"超额回收合计: {excess_total:>14.2f}")
    print(f"总计:         {total:>14.2f}")


# =============================================================================
# Excel 导出函数
# =============================================================================

def export_excel(results, fp, periods, lambda_val=1.1, alpha3=0.85, alpha4=1.15, need_weighted=True):
    """
    将计算结果导出到 Excel 文件：
      四川模式：1.逐时段明细  2.按触发类型汇总  3.常量参数
      福建模式：1.逐日逐时段明细  2.逐日逐时段触发统计  3.按触发类型汇总  4.常量参数

    参数：
        need_weighted — True 表示均价由电量×价格加权算出（四川模式）
                        False 表示均价直接读取（福建模式，逐日逐时段）
    """
    os.makedirs(os.path.dirname(fp), exist_ok=True)
    wb = openpyxl.Workbook()

    # 条件映射：计算 condition -> 输出触发条件
    cond_map = {"缺额回收": "缺额", "超额回收": "超额", "没有偏差": "未触发"}

    # 样式
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    def_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    exc_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    sum_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lef = Alignment(horizontal="left", vertical="center")
    bdr = Border(left=Side(style="thin"), right=Side(style="thin"),
                 top=Side(style="thin"), bottom=Side(style="thin"))

    def _hdr(ws, row, cols):
        for c, v in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = ctr; cell.border = bdr

    def _dat(ws, row, col, val, fmt=None, fill=None, align=ctr):
        cell = ws.cell(row=row, column=col, value=val)
        cell.alignment = align; cell.border = bdr
        if fill: cell.fill = fill
        if fmt: cell.number_format = fmt

    # ========== Sheet 1: 逐时段明细 ==========
    ws1 = wb.active
    ws1.title = "逐时段明细"
    has_day = any(r.get("day") is not None for r in results)

    if need_weighted:
        sheet1_headers = ["时段", "实际用电量汇总(MWh)", "合约电量汇总(MWh)", "比率",
                          "实时市场分时加权均价(元/MWh)", "集中交易分时均价(元/MWh)",
                          "触发条件", "偏差收益回收金额(元)"]
    elif has_day:
        # 福建模式：逐日逐时段
        sheet1_headers = ["日期", "时段", "实际用电量(MWh)", "合约电量(MWh)", "比率",
                          "现货实时市场加权平均价(元/MWh)", "批发侧合同结算均价(元/MWh)",
                          "触发条件", "偏差收益回收金额(元)"]
    else:
        sheet1_headers = ["时段", "实际用电量(MWh)", "合约电量(MWh)", "比率",
                          "现货实时均价(元/MWh)", "集中交易均价(元/MWh)",
                          "触发条件", "偏差收益回收金额(元)"]
    _hdr(ws1, 1, sheet1_headers)

    deficit_total = 0; excess_total = 0
    # 批量构造 Sheet1 数据行（高性能写入）
    sheet1_rows = []
    for r in results:
        tc = cond_map.get(r["condition"], r["condition"])
        fill = def_fill if tc == "缺额" else (exc_fill if tc == "超额" else None)
        if tc == "缺额": deficit_total += r["recovery"]
        elif tc == "超额": excess_total += r["recovery"]
        if has_day:
            sheet1_rows.append([
                f'D{r["day"]:02d}', f'{r["t"]:02d}:00', r["actual"], r["contract"],
                r["ratio"], r["realtime_avg"], r["concentrate_avg"], tc, r["recovery"]
            ])
        else:
            sheet1_rows.append([
                f'{r["t"]:02d}:00', r["actual"], r["contract"], r["ratio"],
                r["realtime_avg"], r["concentrate_avg"], tc, r["recovery"]
            ])
    # 批量追加（openpyxl 原生批量写入，比逐单元格写入快 3-5 倍）
    for row_data in sheet1_rows:
        ws1.append(row_data)
    # 应用样式（批量写入后统一设置格式）
    for i, r in enumerate(results):
        row = i + 2
        tc = cond_map.get(r["condition"], r["condition"])
        fill = def_fill if tc == "缺额" else (exc_fill if tc == "超额" else None)
        col_count = 9 if has_day else 8
        for cc in range(1, col_count + 1):
            cell = ws1.cell(row=row, column=cc)
            cell.alignment = ctr
            cell.border = bdr
            if fill:
                cell.fill = fill
        # 金额列格式化
        amt_col = col_count
        ws1.cell(row=row, column=amt_col).number_format = "#,##0.00"

    hr = len(results) + 2
    col_count = 9 if has_day else 8
    _dat(ws1, hr, 1, "合计", fill=sum_fill)
    for cc in range(2, col_count): _dat(ws1, hr, cc, "", fill=sum_fill)
    _dat(ws1, hr, col_count, round(deficit_total + excess_total, 2), "#,##0.00", fill=sum_fill)
    if has_day:
        for cc, w in enumerate([8, 10, 20, 20, 10, 26, 26, 10, 24], 1):
            ws1.column_dimensions[get_column_letter(cc)].width = w
    else:
        for cc, w in enumerate([10, 22, 22, 10, 28, 28, 10, 24], 1):
            ws1.column_dimensions[get_column_letter(cc)].width = w

    # Sheet 索引：四川模式 Sheet2=按触发类型汇总，福建模式同理
    sheet2_idx = 2

    # ========== Sheet {sheet2_idx}: 按触发类型汇总 ==========
    ws2 = wb.create_sheet("按触发类型汇总")
    if has_day:
        _hdr(ws2, 1, ["触发类型", "触(日×时段)数", "偏差收益回收金额(元)", "占全天比例"])
    else:
        _hdr(ws2, 1, ["触发类型", "触发时段数", "时段列表", "偏差收益回收金额(元)"])

    triggers = {}
    for r in results:
        tc = cond_map.get(r["condition"], r["condition"])
        if tc not in triggers: triggers[tc] = {"count": 0, "periods": [], "recovery": 0}
        triggers[tc]["count"] += 1
        triggers[tc]["periods"].append(f'{r["t"]:02d}:00')
        triggers[tc]["recovery"] += r["recovery"]

    summary_rows = []
    for i, t in enumerate(["缺额", "超额", "未触发"]):
        if t not in triggers: continue
        dd = triggers[t]
        if has_day:
            pct = f'{dd["count"] / len(results) * 100:.1f}%' if len(results) > 0 else '0%'
            summary_rows.append([t, dd["count"], round(dd["recovery"], 2), pct])
        else:
            summary_rows.append([t, dd["count"], ", ".join(dd["periods"]), round(dd["recovery"], 2)])
    for row_data in summary_rows:
        ws2.append(row_data)
    # 统一设置样式
    for i, (t, _) in enumerate([(t, triggers[t]) for t in ["缺额", "超额", "未触发"] if t in triggers]):
        row = i + 2
        tc = t
        fill = def_fill if tc == "缺额" else (exc_fill if tc == "超额" else None)
        col_n = 4 if has_day else 4
        for cc in range(1, col_n + 1):
            cell = ws2.cell(row=row, column=cc)
            cell.alignment = ctr if cc != 3 and has_day else (lef if not has_day and cc == 3 else ctr)
            cell.border = bdr
            if fill:
                cell.fill = fill
        if has_day:
            ws2.cell(row=row, column=3).number_format = "#,##0.00"
        else:
            ws2.cell(row=row, column=4).number_format = "#,##0.00"

    tr = len(triggers) + 2
    _dat(ws2, tr, 1, "合计", fill=sum_fill)
    if has_day:
        _dat(ws2, tr, 2, len(results), fill=sum_fill)
        _dat(ws2, tr, 3, round(deficit_total + excess_total, 2), "#,##0.00", fill=sum_fill)
        _dat(ws2, tr, 4, "100.0%", fill=sum_fill)
        for cc, w in enumerate([10, 16, 24, 14], 1):
            ws2.column_dimensions[get_column_letter(cc)].width = w
    else:
        _dat(ws2, tr, 2, len(results), fill=sum_fill)
        _dat(ws2, tr, 3, "", fill=sum_fill)
        _dat(ws2, tr, 4, round(deficit_total + excess_total, 2), "#,##0.00", fill=sum_fill)
        for cc, w in enumerate([10, 14, 80, 24], 1):
            ws2.column_dimensions[get_column_letter(cc)].width = w

    # ========== Sheet {sheet2_idx+1}: 常量参数 ==========
    ws3 = wb.create_sheet("常量参数")
    _hdr(ws3, 1, ["参数名称", "参数值", "说明"])

    calc_month = _extract_calc_month(fp)
    params = [
        ("计算月份", calc_month, "数据所属月份（从文件名或数据自动提取）"),
        ("λ", lambda_val, "考核系数"),
        ("α₃", f"{alpha3*100:.0f}%", "用电侧缺额偏差触发线"),
        ("α₄", f"{alpha4*100:.0f}%", "用电侧超额偏差触发线"),
        ("超额偏差回收合计(元)", round(excess_total, 2), "触发超额偏差的时段回收金额合计"),
        ("缺额偏差回收合计(元)", round(deficit_total, 2), "触发缺额偏差的时段回收金额合计"),
        ("回收总额(元)", round(deficit_total + excess_total, 2), "超额+缺额合计"),
    ]
    for i, (n, v, desc) in enumerate(params):
        row = i + 2
        _dat(ws3, row, 1, n, align=lef); _dat(ws3, row, 2, v); _dat(ws3, row, 3, desc, align=lef)
    for cc, w in enumerate([24, 18, 40], 1):
        ws3.column_dimensions[get_column_letter(cc)].width = w

    wb.save(fp)
    print()
    print("Excel已导出: " + fp)


# =============================================================================
# 工具函数
# =============================================================================

def _extract_calc_month(file_path):
    """从文件名中提取数据月份（如 202603 → 2026年03月）。"""
    basename = os.path.basename(file_path)
    m = re.search(r'(\d{4})[\-_年]?(\d{2})', basename)
    if m:
        return f"{m.group(1)}年{m.group(2)}月"
    return datetime.date.today().strftime("%Y年%m月")


def extract_province_from_filename(file_path):
    """尝试从文件名中提取省份名。"""
    basename = os.path.basename(file_path)
    name_without_ext = os.path.splitext(basename)[0]

    province_keywords = [
        "四川", "江苏", "浙江", "安徽", "福建", "江西", "山东", "河南",
        "湖北", "湖南", "广东", "广西", "海南", "重庆", "云南", "贵州",
        "陕西", "甘肃", "青海", "宁夏", "新疆", "西藏", "山西", "内蒙",
        "内蒙古", "辽宁", "吉林", "黑龙江", "北京", "上海", "天津", "河北"
    ]

    for kw in province_keywords:
        if kw in name_without_ext:
            return kw

    cleaned = re.sub(
        r'用电侧|中长期|签约|缺额|超额|偏差|收益|回收|计算|数据|输入|模板|样例|示例',
        '', name_without_ext
    )
    cleaned = re.sub(r'[\(\)（）、，。\d\-_]', '', cleaned).strip()
    if cleaned and len(cleaned) <= 6:
        return cleaned

    return None


# =============================================================================
# 主入口
# =============================================================================

if __name__ == "__main__":
    """
    用法：python3 calculate.py <Excel文件路径> [省份] <λ> <α₃> <α₄>

    参数说明：
        Excel文件路径  — 输入数据文件（必填）
        省份           — 省份名称（可选；不传则自动从文件名提取）
        λ              — 考核系数（必填，如 1.1）
        α₃             — 缺额偏差触发线（必填，如 0.85 表示 85%）
        α₄             — 超额偏差触发线（必填，如 1.15 表示 115%）

    示例：
      python3 calculate.py 四川超缺额偏差回收.xlsx 四川 1.1 0.85 1.15
      python3 calculate.py 用电侧中长期签约缺额（超额）偏差收益回收.xlsx 1.1 0.85 1.15
    """
    if len(sys.argv) < 5:
        print("用法: python3 calculate.py <Excel文件路径> [省份] <λ> <α₃> <α₄>")
        print()
        print("必填参数：")
        print("  Excel文件路径  — 输入数据文件")
        print("  λ              — 考核系数（如 1.1）")
        print("  α₃             — 缺额偏差触发线（如 0.85 表示 85%）")
        print("  α₄             — 超额偏差触发线（如 1.15 表示 115%）")
        print()
        print("可选参数：")
        print("  省份           — 省份名称，用于输出文件名；不传则自动从文件名提取")
        print()
        print("示例：")
        print("  python3 calculate.py data.xlsx 四川 1.1 0.85 1.15")
        print("  python3 calculate.py data.xlsx 1.1 0.85 1.15")
        sys.exit(1)

    # 解析命令行参数
    excel_path = sys.argv[1]

    try:
        float(sys.argv[2])
        province = None
        lambda_val = float(sys.argv[2])
        alpha3 = float(sys.argv[3])
        alpha4 = float(sys.argv[4])
    except ValueError:
        province = sys.argv[2]
        lambda_val = float(sys.argv[3])
        alpha3 = float(sys.argv[4])
        alpha4 = float(sys.argv[5])

    if not province:
        province = extract_province_from_filename(excel_path)
    if not province:
        print("错误：无法从文件名自动提取省份名，请通过第二个参数显式传入。")
        print("用法: python3 calculate.py <Excel文件路径> <省份名> <λ> <α₃> <α₄>")
        sys.exit(1)

    # 执行核心计算
    results, periods, need_weighted = calculate(excel_path, lambda_val, alpha3, alpha4)

    # 终端打印
    print_results(results)

    # 导出
    sd = os.path.dirname(os.path.abspath(__file__))
    wd = os.path.dirname(os.path.dirname(os.path.dirname(sd)))
    od = os.path.join(wd, "source", "output", "sc-electricity-excess-deficit-settlement")

    date_str = datetime.date.today().strftime('%Y%m%d')
    out_name = f"{province}超缺额偏差回收{date_str}.xlsx"

    export_excel(results, os.path.join(od, out_name), periods, lambda_val, alpha3, alpha4, need_weighted)
