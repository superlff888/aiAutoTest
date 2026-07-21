#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
aiAutoTester 流水线 - 阶段5：用例设计 + Excel导出
用途：将抽象测试点富化为可执行测试用例，导出为标准Excel

前提：
  - 03_requirements/ 已有各章 requirements.json（requirement-splitter 产出）
  - 05_testpoint_review/ 已有各章 improved_testpoints.json（testpoint-extractor + testpoint-reviewer 产出）
用法：python3 _pipeline.py <项目目录>
例：python3 _pipeline.py .claude/output/aiAutoTester/太乙2.0_支持发电
"""

import sys, os, re, json
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==================== 辅助函数 ====================
def _load_requirements(project_dir):
    """加载所有章节的需求JSON，返回 {req_id: req}"""
    reqs = {}
    req_dir = os.path.join(project_dir, "03_requirements")
    if not os.path.isdir(req_dir):
        return reqs
    for cd in sorted(os.listdir(req_dir)):
        rp = os.path.join(req_dir, cd, "requirements.json")
        if not os.path.exists(rp):
            continue
        with open(rp, "r", encoding="utf-8") as f:
            data = json.load(f)
        for mod in data.get("modules", []):
            for r in mod.get("requirements", []):
                reqs[r["requirement_id"]] = r
    return reqs


# ==================== 用例设计引擎 ====================
# 根据测试点的具体信息（验收标准、需求描述、需求可测试要素）
# 推导具体的测试步骤、预期结果、测试数据

# --- 辅助函数：从文本中提取上下文 ---
def _extract_numbers(text):
    """提取文本中的所有数字（含小数、百分比）"""
    return re.findall(r'[\d]+\.?[\d]*', text)

def _extract_thresholds(text):
    """提取阈值/条件表达式"""
    patterns = [
        r'(?:大[于过]|超过|超出|高于|≥|>=|大于等于)[\s]*(\d+\.?\d*%?)',
        r'(?:小[于于]|低于|少于|≤|<=|小于等于)[\s]*(\d+\.?\d*%?)',
        r'(?:等于|=|为)[\s]*(\d+\.?\d*%?)',
        r'([\d]+\.?[\d*%]*)[\s]*(?:时|的|以|后|才)',
    ]
    results = []
    for p in patterns:
        results.extend(re.findall(p, text))
    return list(set(results))

def _extract_business_objects(text):
    """提取业务对象/实体名称"""
    return re.findall(r'([\u4e00-\u9fa5]+(?:数据|信息|记录|参数|值|金额|电量|电价|价格|用户|角色|配置|规则|条件|公式|系数|阈值|结果|偏差|收益|回收|通知|报告|页面|功能|模块|系统|接口|文件|表格|列表|按钮|菜单|选项|输入|输出|字段))', text)

def _extract_actions(text):
    """提取动作动词"""
    action_kw = ['计算', '校验', '校核', '通知', '推送', '发送', '展示', '显示', '查询',
                  '导出', '导入', '新增', '创建', '删除', '修改', '更新', '提交', '保存',
                  '下载', '上传', '触发', '生成', '返回', '记录', '统计', '汇总', '分摊',
                  '结算', '回收', '发放', '扣除', '返还', '对比', '排序', '筛选', '搜索',
                  '登录', '进入', '点击', '选择', '填写', '确认', '取消', '提交', '刷新']
    found = [a for a in action_kw if a in text]
    return found

def _extract_domain_context(req_desc, tp_name, mapped_criteria):
    """综合提取测试点的业务上下文"""
    combined = f"{req_desc} {tp_name} {mapped_criteria}"
    return {
        'numbers': _extract_numbers(combined),
        'thresholds': _extract_thresholds(combined),
        'business_objects': _extract_business_objects(combined),
        'actions': _extract_actions(combined),
        'formulas': _extract_formulas(req_desc),
        'coefficients': _extract_coefficients(req_desc),
        'examples': _extract_examples(req_desc),
        'has_校核': '校核' in combined or '校验' in combined,
        'has_通知': '通知' in combined or '推送' in combined or '发送' in combined,
        'has_计算': '计算' in combined or '公式' in combined or '结算' in combined or '电费' in combined,
        'has_回收': '回收' in combined or '偏差' in combined or '获利' in combined,
        'has_展示': '展示' in combined or '显示' in combined or '页面' in combined or '报告' in combined,
        'has_排序': '排序' in combined or '顺序' in combined,
        'has_对比': '对比' in combined or '比较' in combined,
        'has_导入导出': '导入' in combined or '导出' in combined or '下载' in combined,
        'has_新增修改删除': any(k in combined for k in ['新增', '创建', '删除', '修改', '更新']),
    }

def _extract_formulas(text):
    """从需求描述中提取计算公式/规则表达式"""
    formulas = []
    for line in text.split('\n'):
        stripped = line.strip()
        if any(kw in stripped for kw in ['公式', '计算', '=', '×', '÷', '+', '-', 'Max(', 'Min(']):
            if len(stripped) > 5 and len(stripped) < 300:
                if any(c in stripped for c in '0123456789.×÷+-='):
                    formulas.append(stripped[:200])
    return formulas

def _extract_coefficients(text):
    """提取系数/参数/默认值"""
    coeffs = []
    patterns = [
        r'([\u4e00-\u9fa5]+[系数阈值参数比例费率倍数因子])\s*[=为:]\s*([\d]+\.?[\d*%]*)',
        r'([\u4e00-\u9fa5]*[系数阈值参数比例费率])\s*(?:默认值|取值为?|为|是)\s*([\d]+\.?[\d*%]*)',
        r'([\u4e00-\u9fa5]+[值量额价率比])\s*<\s*([\d]+\.?[\d*%]*)',
        r'([\u4e00-\u9fa5]+[值量额价率比])\s*>\s*([\d]+\.?[\d*%]*)',
        r'([\u4e00-\u9fa5]+[值量额价率比])\s*[=为]\s*([\d]+\.?[\d*%]*)',
    ]
    for p in patterns:
        for m in re.findall(p, text):
            item = f"{m[0]}={m[1]}"
            if item not in coeffs:
                coeffs.append(item)
    return coeffs

def _extract_examples(text):
    """提取算例/示例中的具体数据"""
    examples = []
    for line in text.split('\n'):
        stripped = line.strip()
        if any(kw in stripped for kw in ['算例', '例如', '示例', '如：', '如:', '比如']):
            if len(stripped) > 5 and len(stripped) < 400:
                nums = re.findall(r'[\d]+\.?[\d]*', stripped)
                if nums and len(stripped) > 10:
                    examples.append(stripped[:200])
    return examples

def _derive_ft_steps(ctx, mapped_criteria, req_name, chapter_title):
    """根据上下文推导 FT 测试步骤"""
    steps = []
    if ctx['has_展示']:
        steps.append(f"1. 登录系统，进入「{chapter_title}」页面")
    elif ctx['has_计算'] or ctx['has_回收']:
        steps.append(f"1. 登录系统，进入「{chapter_title}」计算/校核页面")
    elif ctx['has_通知']:
        steps.append(f"1. 登录系统，进入「{chapter_title}」通知管理页面")
    else:
        steps.append(f"1. 登录系统，进入「{chapter_title}」相关功能页面")

    if ctx['business_objects']:
        obj = ctx['business_objects'][0]
        steps.append(f"2. 准备{obj}：根据验收标准要求设置输入条件")
    else:
        steps.append(f"2. 准备{req_name}所需的输入数据")

    if ctx['actions']:
        action = ctx['actions'][0]
        steps.append(f"3. 执行{action}操作，触发{req_name}功能")
    else:
        steps.append(f"3. 执行{req_name}相关操作")

    if mapped_criteria:
        steps.append(f"4. 验证结果：{mapped_criteria[:50]}")
    else:
        steps.append(f"4. 验证{req_name}功能输出符合预期")

    if ctx['thresholds']:
        th = ctx['thresholds'][0]
        steps.append(f"5. 确认阈值条件（{th}）判定正确，边界值处理符合规则")
    elif ctx['numbers']:
        nums = ctx['numbers'][:3]
        steps.append(f"5. 确认关键数值（{', '.join(nums)}）计算/处理正确")

    return steps

def _derive_ft_expected(ctx, mapped_criteria, req_name):
    """根据上下文推导 FT 预期结果"""
    expected = []
    if mapped_criteria:
        expected.append(mapped_criteria[:80])
    if ctx['has_计算']:
        expected.append(f"{req_name}计算结果准确，与手工验算一致")
    if ctx['has_校核']:
        expected.append("校核判定正确，通过/不通过判定准确")
    if ctx['has_通知']:
        expected.append("通知发送及时，内容准确无误")
    if ctx['has_展示']:
        expected.append("页面数据展示完整，格式正确")
    if not expected:
        expected.append(f"{req_name}功能执行正常，结果符合需求描述")
    expected.append("页面/接口无异常报错")
    return expected

def _derive_bt_steps(ctx, req_name, chapter_title):
    """根据上下文推导 BT 测试步骤"""
    steps = []
    steps.append(f"1. 登录系统，进入「{chapter_title}」相关页面")

    if ctx['thresholds']:
        min_val = ctx['thresholds'][0]
        steps.append(f"2. 场景A-下边界：输入最小有效值（{min_val}），验证系统处理")
    else:
        steps.append(f"2. 场景A-下边界：输入{req_name}的最小有效值（0或最小允许值），验证系统处理")

    if ctx['thresholds'] and len(ctx['thresholds']) > 1:
        max_val = ctx['thresholds'][1]
        steps.append(f"3. 场景B-上边界：输入最大有效值（{max_val}），验证系统处理")
    else:
        steps.append(f"3. 场景B-上边界：输入{req_name}的最大允许值（如999999.99），验证系统处理")

    if ctx['thresholds']:
        th = ctx['thresholds'][0]
        steps.append(f"4. 场景C-临界值：输入刚好等于阈值（{th}）的值，验证判定逻辑")
    else:
        steps.append(f"4. 场景C-临界值：输入等于阈值的值，验证{req_name}判定逻辑正确")

    if ctx['thresholds']:
        th = ctx['thresholds'][0]
        steps.append(f"5. 场景D-越界值：输入超出阈值±0.01的值，验证拦截或告警")
    else:
        steps.append(f"5. 场景D-越界值：输入超出允许范围±0.01的值，验证系统拦截")

    steps.append(f"6. 场景E-空值：不输入{req_name}所需数据（空字符串/NULL/空列表），验证处理")
    steps.append(f"7. 分别记录各场景的系统响应和输出结果")

    return steps

def _derive_bt_expected(ctx, req_name):
    """根据上下文推导 BT 预期结果"""
    expected = []
    if ctx['has_计算']:
        expected.extend([
            "场景A/B/C：边界值计算结果准确，不溢出",
            "场景D：越界值被拦截或触发告警提示",
            "场景E：空值输入时提示'数据不能为空'或按默认值处理",
        ])
    elif ctx['has_校核']:
        expected.extend([
            "场景A/B/C：校核判定正确，边界值校核通过",
            "场景D：越界值校核不通过，触发告警",
            "场景E：空值校核时提示'数据为空，无法校核'",
        ])
    else:
        expected.extend([
            "场景A/B/C：边界条件下功能正常执行，结果正确",
            "场景D：越界值被正确拦截，有明确提示",
            "场景E：空值处理合理，不崩溃",
        ])
    expected.append("所有场景系统不崩溃，有明确的提示信息")
    return expected

def _derive_et_steps(ctx, req_name, chapter_title):
    """根据上下文推导 ET 测试步骤"""
    steps = []
    steps.append(f"1. 登录系统，进入「{chapter_title}」相关页面")

    if ctx['has_计算']:
        steps.extend([
            f"2. 场景A-数据源异常：{req_name}接口返回空数据/null，验证系统处理",
            f"3. 场景B-超时异常：模拟{req_name}接口超时（>30s），验证超时处理",
            f"4. 场景C-格式异常：输入非法格式数据（如字母代替数字），验证校验拦截",
            f"5. 场景D-并发异常：同时发起2次{req_name}请求，验证结果一致性",
            f"6. 场景E-权限异常：无权限用户尝试执行{req_name}，验证权限拦截",
        ])
    elif ctx['has_通知']:
        steps.extend([
            f"2. 场景A-接口异常：模拟通知推送接口返回500错误，验证降级处理",
            f"3. 场景B-超时异常：通知接口超时30s未响应，验证超时处理",
            f"4. 场景C-数据异常：通知内容超长或包含特殊字符，验证截断处理",
            f"5. 场景D-高频触发：1分钟内连续触发10次，验证限流处理",
            f"6. 场景E-权限异常：无权限用户查看/发送通知，验证权限拦截",
        ])
    else:
        steps.extend([
            f"2. 场景A-接口异常：{req_name}功能接口返回500错误，验证错误提示",
            f"3. 场景B-超时异常：操作超时30秒未响应，验证超时提示",
            f"4. 场景C-数据异常：输入非法格式数据，验证校验拦截",
            f"5. 场景D-并发异常：同时执行相同操作2次，验证幂等处理",
            f"6. 场景E-权限异常：无权限用户尝试操作，验证权限拦截",
        ])

    steps.append(f"7. 分别记录各场景的系统响应和错误信息")
    return steps

def _derive_et_expected(ctx, req_name):
    """根据上下文推导 ET 预期结果"""
    expected = []
    if ctx['has_计算']:
        expected.extend([
            "场景A：接口返回空数据时提示'暂无数据'，不崩溃",
            "场景B：超时后提示'计算超时，请稍后重试'，页面可继续操作",
            "场景C：非法格式数据被拦截，提示'数据格式错误'",
            "场景D：并发请求仅执行1次，结果一致",
            "场景E：无权限用户提示'您无权执行这操作'",
        ])
    elif ctx['has_通知']:
        expected.extend([
            "场景A：接口500时记录失败日志，提示'通知发送失败'",
            "场景B：超时后提示'通知发送超时'，不阻断其他功能",
            "场景C：超长内容自动截断，消息发送成功",
            "场景D：高频触发时仅发送前N条，后续提示'操作过于频繁'",
            "场景E：无权限用户提示'您无权发送通知'",
        ])
    else:
        expected.extend([
            "场景A：接口500时显示'系统异常，请稍后重试'，不白屏",
            "场景B：超时后提示'操作超时，请重试'，页面可继续操作",
            "场景C：非法格式被拦截，提示'输入格式不正确'",
            "场景D：并发操作仅执行1次，结果一致",
            "场景E：无权限提示'您无权执行这操作'",
        ])
    expected.append("所有异常场景有明确的错误提示和恢复路径")
    return expected

def _derive_test_data(ctx, tp_type, mapped_criteria, req_name):
    """根据上下文推导测试数据，优先使用需求文档中的公式/规则/数值"""
    data = []

    if tp_type == "FT":
        if ctx['formulas']:
            for i, formula in enumerate(ctx['formulas'][:3], 1):
                data.append({f"公式/规则{i}": formula[:100]})
        if ctx['coefficients']:
            for c in ctx['coefficients'][:4]:
                data.append({"参数": c[:80]})
        if ctx['examples']:
            for i, ex in enumerate(ctx['examples'][:2], 1):
                data.append({f"算例{i}": ex[:100]})
        if ctx['thresholds']:
            for th in ctx['thresholds'][:2]:
                data.append({"阈值条件": th})
        if ctx['business_objects']:
            for obj in ctx['business_objects'][:2]:
                data.append({"测试对象": obj})
        if mapped_criteria:
            data.append({"验收标准": mapped_criteria[:60]})
        if not data:
            data.append({"测试对象": req_name, "验证点": "功能执行结果符合预期"})

    elif tp_type == "BT":
        if ctx['thresholds']:
            th = ctx['thresholds'][0]
            data.extend([
                {"场景A-下边界": f"输入值 = {th}（最小有效值）"},
                {"场景B-上边界": f"输入值 = 系统支持的最大值"},
                {"场景C-临界值": f"输入值 = {th}（刚好等于阈值）"},
                {"场景D-越界值": f"输入值 = {th} ± 0.01"},
                {"场景E-空值": "输入值 = None 或 '' 或 []"},
            ])
        elif ctx['coefficients']:
            c = ctx['coefficients'][0]
            data.extend([
                {"场景A-下边界": f"输入值 = {c}（最小有效值）"},
                {"场景B-上边界": "输入值 = 系统支持的最大值"},
                {"场景C-临界值": f"输入值 = {c}（临界值）"},
                {"场景D-越界值": f"输入值 = 临界值 ± 0.01"},
                {"场景E-空值": "输入值 = None 或 '' 或 []"},
            ])
        else:
            data.extend([
                {"场景A-下边界": "输入值 = 0 或最小允许值"},
                {"场景B-上边界": "输入值 = 999999.99"},
                {"场景C-临界值": "输入值 = 刚好等于阈值的值"},
                {"场景D-越界值": "输入值 = 阈值 ± 0.01"},
                {"场景E-空值": "输入值 = None 或 '' 或 []"},
            ])
        if ctx['formulas']:
            data.append({"相关公式": ctx['formulas'][0][:80]})

    elif tp_type == "ET":
        data.extend([
            {"场景A-接口异常": "模拟接口返回 HTTP 500 / null"},
            {"场景B-超时": "模拟接口响应时间 > 30000ms"},
            {"场景C-格式异常": "输入: 'abc' (预期为数字) 或 {} 缺少必需字段"},
            {"场景D-并发": "同时发起2次相同请求，间隔 < 100ms"},
            {"场景E-权限": "使用普通用户账号（非管理员）尝试操作"},
        ])

    elif tp_type == "UIT":
        data.append({"对比基准": "需求文档中的关联截图"})

    return data


def _enrich_testpoint(tp, chapter_title, reqs):
    """用例设计：根据测试点的具体信息推导可执行的测试用例"""
    req_id = tp.get("requirement_id", "")
    req_name = tp.get("requirement_name", "")
    tp_name = tp.get("testpoint_name", "")
    tp_type = tp.get("test_type", "FT")
    req = reqs.get(req_id, {})
    req_desc = req.get("detailed_description", "")
    mapped_criteria = tp.get("mapped_criteria", "")

    ctx = _extract_domain_context(req_desc, tp_name, mapped_criteria)

    if tp_type == "FT":
        tp["test_steps"] = _derive_ft_steps(ctx, mapped_criteria, req_name, chapter_title)
        tp["expected_results"] = _derive_ft_expected(ctx, mapped_criteria, req_name)
        tp["test_data"] = _derive_test_data(ctx, "FT", mapped_criteria, req_name)
        tp["notes"] = f"功能测试"

    elif tp_type == "BT":
        tp["test_steps"] = _derive_bt_steps(ctx, req_name, chapter_title)
        tp["expected_results"] = _derive_bt_expected(ctx, req_name)
        tp["test_data"] = _derive_test_data(ctx, "BT", mapped_criteria, req_name)
        tp["notes"] = "边界测试"

    elif tp_type == "ET":
        tp["test_steps"] = _derive_et_steps(ctx, req_name, chapter_title)
        tp["expected_results"] = _derive_et_expected(ctx, req_name)
        tp["test_data"] = _derive_test_data(ctx, "ET", mapped_criteria, req_name)
        tp["notes"] = "异常测试"

    elif tp_type == "UIT":
        tp["test_data"] = _derive_test_data(ctx, "UIT", mapped_criteria, req_name)
        tp["notes"] = "UI测试：对照需求文档中的截图验证"

    return tp


# ==================== 阶段5：导出Excel ====================
def export_excel(project_dir):
    """读取 improved_testpoints.json，富化用例并导出Excel"""
    review_base = os.path.join(project_dir, "05_testpoint_review")
    reqs = _load_requirements(project_dir)
    export_base = os.path.join(project_dir, "06_testcase_export")
    os.makedirs(export_base, exist_ok=True)

    # 样式
    title_font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="微软雅黑", size=9)
    title_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    columns = ["需求ID", "需求名称", "用例ID", "用例名称", "优先级", "测试类型", "前置条件", "测试步骤", "预期结果", "备注", "标签", "维护人", "测试数据", "评审状态"]
    type_map = {"FT": "功能测试", "BT": "边界测试", "ET": "异常测试", "PT": "性能测试",
                "UIT": "UI测试", "ST": "安全测试", "CT": "合规测试", "DT": "数据测试"}
    priority_map = {"P0": "P0-重要紧急", "P1": "P1-重要不紧急", "P2": "P2-紧急不重要", "P3": "P3-不紧急不重要"}

    wb = openpyxl.Workbook()
    all_tps = []
    chapter_names = []

    for d in sorted(os.listdir(review_base)):
        tp_path = os.path.join(review_base, d, "improved_testpoints.json")
        if not os.path.isfile(tp_path): continue
        with open(tp_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        testpoints = data.get("testpoints", [])
        if not testpoints: continue

        chapter_title = data.get("chapter_title", d)

        # 用例设计：将抽象测试点富化为可执行的测试用例
        for tp in testpoints:
            _enrich_testpoint(tp, chapter_title, reqs)

        chapter_names.append(d)
        ws = wb.create_sheet(title=d)

        # 标题
        last_col = get_column_letter(len(columns))
        ws.merge_cells(f"A1:{last_col}1")
        cell = ws["A1"]
        cell.value = f"测试用例 - {d}"
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 35

        # 表头
        for ci, cn in enumerate(columns, 1):
            c = ws.cell(row=2, column=ci, value=cn)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
        ws.row_dimensions[2].height = 25

        # 数据
        for ri, tp in enumerate(testpoints, 3):
            steps = tp.get("test_steps", [])
            fixed_steps = []
            for i, s in enumerate(steps, 1):
                cleaned = re.sub(r"^\d+\.\s*", "", s)
                fixed_steps.append(f"{i}. {cleaned}")

            td = tp.get("test_data", [])
            if isinstance(td, list):
                td_str = "; ".join(
                    ", ".join(f"{k}: {v}" for k, v in item.items()) if isinstance(item, dict) else str(item)
                    for item in td
                )
            else:
                td_str = str(td)

            row = [
                tp.get("requirement_id", ""),
                tp.get("requirement_name", ""),
                tp.get("testpoint_id", ""),
                tp.get("testpoint_name", ""),
                priority_map.get(tp.get("priority", "P1"), tp.get("priority", "P1")),
                type_map.get(tp.get("test_type", "FT"), tp.get("test_type", "FT")),
                "、".join(tp.get("preconditions", [])),
                "\n".join(fixed_steps),
                "\n".join(tp.get("expected_results", [])),
                tp.get("notes", ""),
                "AI用例",
                "李房房-38926763",
                td_str,
                "已评审"
            ]
            for ci, val in enumerate(row, 1):
                c = ws.cell(row=ri, column=ci, value=str(val))
                c.font = body_font
                c.border = thin_border
                c.alignment = Alignment(vertical="top", wrap_text=True)
            if ri % 2 == 0:
                for ci in range(1, len(columns) + 1):
                    ws.cell(row=ri, column=ci).fill = alt_fill
            ws.row_dimensions[ri].height = 60

        col_widths = [18, 22, 22, 40, 8, 10, 25, 50, 45, 30, 10, 18, 30, 10]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        all_tps.extend(testpoints)
        print(f"  {d}: {len(testpoints)} 条用例")

    # 统计汇总 Sheet
    ws_s = wb.create_sheet(title="统计汇总", index=0)
    ws_s.merge_cells("A1:B1")
    ws_s["A1"].value = "测试用例统计汇总"
    ws_s["A1"].font = title_font
    ws_s["A1"].fill = title_fill
    ws_s["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_s.row_dimensions[1].height = 35

    stats = [
        ("项目", os.path.basename(project_dir)),
        ("总测试用例数", str(len(all_tps))),
        ("", ""),
        ("按类型统计", ""),
        ("功能测试 (FT)", str(sum(1 for t in all_tps if t.get("test_type") == "FT"))),
        ("边界测试 (BT)", str(sum(1 for t in all_tps if t.get("test_type") == "BT"))),
        ("异常测试 (ET)", str(sum(1 for t in all_tps if t.get("test_type") == "ET"))),
        ("UI测试 (UIT)", str(sum(1 for t in all_tps if t.get("test_type") == "UIT"))),
        ("", ""),
        ("按优先级统计", ""),
        ("P0-重要紧急", str(sum(1 for t in all_tps if t.get("priority") == "P0"))),
        ("P1-重要不紧急", str(sum(1 for t in all_tps if t.get("priority") == "P1"))),
        ("P2-紧急不重要", str(sum(1 for t in all_tps if t.get("priority") == "P2"))),
        ("P3-不紧急不重要", str(sum(1 for t in all_tps if t.get("priority") == "P3"))),
        ("", ""),
        ("按章节统计", ""),
    ]
    for cn in chapter_names:
        ctps = sum(1 for t in all_tps if t.get("testpoint_id", "").startswith(cn[:4].replace("第", "").replace("_", "")))
        stats.append((cn, str(ctps)))

    for ri, (k, v) in enumerate(stats, 3):
        ws_s.cell(row=ri, column=1, value=k).font = body_font
        ws_s.cell(row=ri, column=2, value=v).font = body_font
    ws_s.column_dimensions["A"].width = 30
    ws_s.column_dimensions["B"].width = 20

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    today = datetime.now().strftime("%Y%m%d")
    filename = f"{os.path.basename(project_dir)}_测试用例_{today}.xlsx"
    filepath = os.path.join(export_base, filename)
    wb.save(filepath)
    print(f"\n[OK] Excel 已保存: {filepath}")
    print(f"   总测试用例数: {len(all_tps)}")
    return filepath


# ==================== 主流程 ====================
def main():
    if len(sys.argv) < 2:
        print("用法: python3 _pipeline.py <项目目录>")
        print("例: python3 _pipeline.py .claude/output/aiAutoTester/太乙2.0_支持发电")
        print()
        print("前提: 项目目录下需有 05_testpoint_review/ 各章 improved_testpoints.json")
        print("     （由 testpoint-extractor + testpoint-reviewer 技能产出）")
        sys.exit(1)

    project_dir = sys.argv[1]

    # 验证前置条件
    review_dir = os.path.join(project_dir, "05_testpoint_review")
    if not os.path.isdir(review_dir):
        print(f"[ERROR] 未找到 {review_dir}")
        print("       请先执行 testpoint-extractor + testpoint-reviewer 技能生成 improved_testpoints.json")
        sys.exit(1)

    print("阶段5: 用例设计 + 导出 Excel")
    filepath = export_excel(project_dir)

    print(f"\n[DONE] 导出完成！")


if __name__ == "__main__":
    main()
