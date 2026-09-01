#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""组合稳定性分布 · 指标计算器数据写入工具

从 algorithmRetrospectiveCombinations.json 报文中提取指定模型组合的数据，
按需求口径计算每日排名，收益按报文原值(元)写入指标计算器 Excel 的 A16:E47 数据区：
  A列=日期  B列=每日排名r  C列=策略收益(元)  D列=实际收益(元)  E列=有效天标志(公式重建)
  C/D列写报文原值(单位:元，显示2位小数，存储原值精度保证C>D胜天判定不受舍入影响)

口径基准（用户确认 + spot_review.js 代码验证）：
  有效天   = 当日策略收益≠0且非空 且 该日实际收益数据存在
             （考核数据与实际收益数据是两套数据，但存在性判定等价，以实际收益是否存在判定）
  每日排名 = 竞赛排名法：当日参与组合按策略收益降序，并列同名次、下一名次跳过
             （名次=当日收益严格更高者个数+1，对齐报文后端dailyRank口径；2026-08-26由相邻名次法改入）
             参与口径与有效天一致（需求口径）；--rank-mode frontend 复现 847 行缺陷行为
  覆盖率   = validDays ÷ 区间天数D，D=len(dailyDataList)，不可用 len(actualBaseline)
  胜率     = winDays ÷ validDays，胜天 = 当日策略收益 > 当日实际收益（跑赢实际）
  综合评价标签 = 全量组合LCB升序分位定档（对齐spot_review.js:1070-1097）：
             LCB=μ+z·σ/√T(z=1,对齐879行)；≤P10稳定优秀/≤P35较稳定/>P90易爆雷/>P65不稳定/else表现中等

布局（2026-08-21版，用户在16/17行插入两行后）：
  A5=指标计算区标题(合并A5:E5锚点)，动态显示目标组合：指标计算区(<label>)
  B16=综合评价标签值  B17=分位线与值("P10(63.8)/P35(78.5)/P65(93.9)/P90(118.2)")
  18行=数据表头  19~49行=数据区(A日期/B排名/C策略收益元/D实际收益元/E有效天标志)
  A51=口径说明标题  A52:E59=合并说明区
"""
import argparse
import json
import math
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

DATA_START = 19          # 数据区起始行（18行为表头，16/17行为综合评价标签/分位线）
DATA_MAX_ROWS = 31       # 计算器容量（19~49）
TITLE_CELL = 'A5'        # 指标计算区标题(合并区A5:E5锚点，动态显示目标组合)
TAG_CELL = 'B16'         # 综合评价标签
QUANT_CELL = 'B17'       # 分位线与值
NOTE_TITLE_CELL = 'A51'  # 口径说明标题
INVALID_FILL = PatternFill('solid', fgColor='FFF2CC')
NO_FILL = PatternFill(fill_type=None)
E_FORMULA = '=IF(AND(B{r}<>"",C{r}<>"",C{r}<>0,D{r}<>""),1,0)'


def find_project_root() -> Path:
    """从脚本位置向上找含 .qoder 目录的祖先作为项目根。"""
    for anc in Path(__file__).resolve().parents:
        if (anc / '.qoder').is_dir():
            return anc
    return Path.cwd()


def sp_of(combo, dt):
    """组合在指定日期的 strategyProfit；键缺失或值为 null 均返回 None。"""
    return next((dd.get('strategyProfit') for dd in combo['dailyDataList'] if dd['date'] == dt), None)


def daily_ranks(combos, base, ti, dates, mode):
    """计算目标组合每日排名。返回 {date: rank}。

    requirement：当日策略收益非零非空 且 当日有实际收益才参与（需求口径=有效天口径）
    frontend   ：仅策略收益非零非空即参与（复现 spot_review.js:847 缺陷，不查实际收益）
    排名规则为竞赛排名法（对齐报文后端dailyRank）：并列同名次、下一名次跳过，
    即名次 = 当日收益严格更高的参与组合个数 + 1。
    """
    ranks = {}
    for dt in dates:
        spt = sp_of(combos[ti], dt)
        if spt in (None, 0):
            continue
        if mode == 'requirement' and dt not in base:
            continue
        part = [(i, sp_of(c, dt)) for i, c in enumerate(combos) if sp_of(c, dt) not in (None, 0)]
        ranks[dt] = sum(1 for _, v in part if v > spt) + 1
    return ranks


def compute_lcb_tags(combos, base, mode, ti):
    """全量组合计算LCB并按分位定档（对齐spot_review.js:860-880, 1070-1097）。

    返回 (tag, quant_text, tgt_lcb)：目标组合(ti)标签、分位线文本、目标LCB。
    排名矩阵按 mode 口径：requirement=有效天参与（需求）；frontend=仅查收益（847行缺陷）。
    """
    dates = [dd['date'] for dd in combos[0]['dailyDataList']]
    sp = [{dd['date']: dd.get('strategyProfit') for dd in c['dailyDataList']} for c in combos]
    rm = [[None] * len(dates) for _ in combos]
    for di, dt in enumerate(dates):
        if mode == 'requirement' and dt not in base:
            continue
        part = [(i, sp[i][dt]) for i in range(len(combos)) if sp[i].get(dt) not in (None, 0)]
        part.sort(key=lambda x: (-x[1], x[0]))
        pos = 0                                              # 竞赛排名法：并列同名次、跳号
        while pos < len(part):
            j = pos
            while j + 1 < len(part) and part[j + 1][1] == part[pos][1]:
                j += 1
            for k in range(pos, j + 1):
                rm[part[k][0]][di] = pos + 1
            pos = j + 1
    lcbs, tgt = [], None
    for i, c in enumerate(combos):
        rs = [r for r in rm[i] if r is not None]
        if not rs:
            continue
        mu = sum(rs) / len(rs)
        sd = math.sqrt(sum((r - mu) ** 2 for r in rs) / len(rs))
        lcb = round((mu + sd / math.sqrt(len(rs))) * 10) / 10
        lcbs.append(lcb)
        if i == ti:
            tgt = lcb
    s = sorted(lcbs)
    n = len(s)
    q = lambda f: s[min(n - 1, math.floor(n * f))]
    p10, p35, p65, p90 = q(0.1), q(0.35), q(0.65), q(0.9)
    tag = ('稳定优秀' if tgt <= p10 else '较稳定' if tgt <= p35
           else '易爆雷' if tgt > p90 else '不稳定' if tgt > p65 else '表现中等')
    quant = f'P10({p10})/P35({p35})/P65({p65})/P90({p90})'
    return tag, quant, tgt


def invalid_reason(combo, base, dt):
    """无效天归因：None=有效天；否则返回类型描述。"""
    sp = sp_of(combo, dt)
    if sp is None:
        return '策略收益键缺失'
    if sp == 0:
        return '策略收益=0'
    if dt not in base:
        return '实际收益缺失'
    return None


def main():
    ap = argparse.ArgumentParser(description='组合稳定性分布·指标计算器数据写入')
    ap.add_argument('--json', default=None, help='报文路径')
    ap.add_argument('--excel', default=None, help='计算器Excel路径')
    ap.add_argument('--combo', default=None, help='目标组合label，缺省自动选validDays最高者')
    ap.add_argument('--rank-mode', choices=['requirement', 'frontend'], default='requirement',
                    help='排名口径：requirement=有效天参与(默认)；frontend=复现前端847行缺陷')
    ap.add_argument('--list', action='store_true', help='仅列出组合摘要，不写Excel')
    ap.add_argument('--dry-run', action='store_true', help='只计算与校验，不写Excel')
    ap.add_argument('--top', type=int, default=20, help='--list显示条数')
    args = ap.parse_args()

    root = find_project_root()
    json_path = Path(args.json) if args.json else root / '.qoder' / 'skills' / 'combo-stability-writer' / 'algorithmRetrospectiveCombinations.json'
    excel_path = Path(args.excel) if args.excel else root / '.qoder' / 'skills' / 'combo-stability-writer' / '组合稳定性分布_指标计算器.xlsx'

    # ---- 1. 读取报文 ----
    try:
        payload = json.loads(json_path.read_text(encoding='utf-8'))
    except FileNotFoundError:
        sys.exit(f'[ERROR] 报文不存在: {json_path}')
    model = payload.get('model') or payload
    combos = model.get('combinations') or []
    base = {x['date']: x['actualProfit'] for x in model.get('actualBaseline') or []}
    if not combos:
        sys.exit('[ERROR] 报文中无 combinations 数据')

    # ---- 2. --list 模式 ----
    if args.list:
        rows = sorted(enumerate(combos), key=lambda x: -x[1]['validDays'])[:args.top]
        print(f'共 {len(combos)} 个组合（按 validDays 降序前 {len(rows)}）：')
        print(f"{'validDays':>9} {'winDays':>7} {'winRate':>8} {'coverage':>9}  label")
        for _, c in rows:
            print(f"{c['validDays']:>9} {c['winDays']:>7} {c['winRate']:>8.4f} {c['coverage']:>9.4f}  {c['label']}")
        return

    # ---- 3. 选定目标组合 ----
    if args.combo:
        ti = next((i for i, c in enumerate(combos) if c['label'] == args.combo), None)
        if ti is None:
            sys.exit(f"[ERROR] 未找到组合: {args.combo}\n提示: 用 --list 查看可用组合 label")
    else:
        ti = max(range(len(combos)), key=lambda i: combos[i]['validDays'])
    target = combos[ti]
    dates = [dd['date'] for dd in target['dailyDataList']]
    n = len(dates)
    if n > DATA_MAX_ROWS:
        sys.exit(f'[ERROR] dailyDataList 长度 {n} 超过计算器容量 {DATA_MAX_ROWS} 行')

    # ---- 4. 计算排名、指标与综合评价标签 ----
    ranks = daily_ranks(combos, base, ti, dates, args.rank_mode)
    valid_days = [dt for dt in dates if invalid_reason(target, base, dt) is None]
    t = len(valid_days)
    w = sum(1 for dt in valid_days if sp_of(target, dt) > base[dt])
    win_rate, coverage = (w / t if t else 0), (t / n if n else 0)
    tag, quant_text, tgt_lcb = compute_lcb_tags(combos, base, args.rank_mode, ti)

    # ---- 5. 与报文交叉校验 ----
    print('=' * 62)
    print(f"目标组合: {target['label']}")
    print(f"区间: {dates[0]} ~ {dates[-1]}  D=len(dailyDataList)={n}  A=len(actualBaseline)={len(base)}")
    print(f"排名口径: {'requirement(有效天参与)' if args.rank_mode == 'requirement' else 'frontend(复现847行缺陷,仅B列对比)'}")
    print('-' * 62)
    checks = [
        ('有效天数 T', t, target['validDays'], 0),
        ('胜天数 W', w, target['winDays'], 0),
        ('胜率', round(win_rate, 4), round(target['winRate'], 4), 5e-5),
        ('覆盖率', round(coverage, 4), round(target['coverage'], 4), 5e-5),
    ]
    ok = True
    for name, got, exp, tol in checks:
        match = abs(got - exp) <= tol
        ok &= match
        print(f"  {name:<8} 计算={got:<10} 报文={exp:<10} {'PASS' if match else 'FAIL'}")
    if not ok:
        sys.exit('[ERROR] 交叉校验失败，终止写入（请检查报文口径是否变化）')

    rs = [ranks[dt] for dt in valid_days if dt in ranks]
    if rs:
        mu = sum(rs) / len(rs)
        sd = math.sqrt(sum((x - mu) ** 2 for x in rs) / len(rs))
        topk = max(1, math.ceil(len(combos) * 0.1))
        print(f"  预计算(报文无此字段): μ={mu:.1f} σ={sd:.1f} 前10%率={sum(1 for x in rs if x <= topk)/len(rs)*100:.1f}% "
              f"爆雷率={sum(1 for x in rs if x > len(combos)-topk)/len(rs)*100:.1f}% LCB={mu+sd/math.sqrt(len(rs)):.1f}")
    invalids = [(dt, invalid_reason(target, base, dt)) for dt in dates if invalid_reason(target, base, dt)]
    print(f"  无效天 {len(invalids)} 个: " + ('; '.join(f'{d}({r})' for d, r in invalids) if invalids else '无'))
    print(f"  综合评价标签: {tag} | LCB={tgt_lcb} | 分位线 {quant_text}")
    print('=' * 62)
    if args.dry_run:
        print('[DRY-RUN] 校验通过，未写入 Excel')
        return

    # ---- 6. 写入 Excel ----
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        sys.exit(f'[ERROR] 计算器不存在: {excel_path}')
    ws = wb.active
    for i, dt in enumerate(dates):
        r = DATA_START + i
        sp, ap_val = sp_of(target, dt), base.get(dt)
        ws.cell(row=r, column=1).value = dt                       # A 日期
        ws.cell(row=r, column=2).value = ranks.get(dt)            # B 每日排名(无效天=None清空)
        c3 = ws.cell(row=r, column=3)                             # C 策略收益(元)
        c3.value = sp                                             # 报文原值直接写入(单位:元)
        c3.number_format = '0.00'
        c4 = ws.cell(row=r, column=4)                             # D 实际收益(元)
        c4.value = ap_val                                         # 报文原值直接写入(单位:元)
        c4.number_format = '0.00'
    # 清空超出 n 的旧行（openpyxl 陷阱：cell(value=None) 不清空，必须 .value=None）
    for r in range(DATA_START + n, DATA_START + DATA_MAX_ROWS):
        for col in range(1, 5):
            ws.cell(row=r, column=col).value = None
        ws.cell(row=r, column=3).number_format = '0.00'
        ws.cell(row=r, column=4).number_format = '0.00'
    # E 列公式重建 + 高亮重置（先全清，再对无效天行高亮）
    for r in range(DATA_START, DATA_START + DATA_MAX_ROWS):
        ws.cell(row=r, column=5).value = E_FORMULA.format(r=r)
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = NO_FILL
    hl_rows = [DATA_START + i for i, dt in enumerate(dates) if invalid_reason(target, base, dt)]
    for r in hl_rows:
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = INVALID_FILL
    ws['B4'] = n                                                  # 区间天数 D 同步
    ws['C4'] = '← 覆盖率的分母'                                   # 校正注释(D仅为覆盖率分母，爆雷率分母是T)
    ws['B2'] = len(combos)                                        # 组合总数(label个数)同步
    title_text = f"指标计算区({target['label']})"                  # 计算区标题 → A5(随组合切换)
    ws[TITLE_CELL] = title_text
    ws[TAG_CELL] = tag                                            # 综合评价标签 → B16
    ws[QUANT_CELL] = quant_text                                   # 分位线与值 → B17
    # 口径说明区动态更新（兼容两种模板：A52-A59逐行独立 / A52:E59整体合并）
    line3 = ('3. 无效天示例(高亮行%s)：%s' % (
        f'{hl_rows[0]}-{hl_rows[-1]}' if len(hl_rows) > 1 else (hl_rows[0] if hl_rows else '无'),
        '；'.join(f'{d}({r})' for d, r in invalids) if invalids else '本组合无无效天'))
    line4 = (f'4. 区间天数D(B4)=len(dailyDataList)={n}，覆盖率分母=D；本表有实际收益的天数A={len(base)}'
             + (f'，若误用A做分母覆盖率虚高为{t}/{len(base)}={t/len(base)*100:.1f}%(正确值{coverage*100:.1f}%)'
                if len(base) < n else '，本区间D=A无缺实际收益日'))
    line6 = ('6. σ为总体标准差STDEV.P；μ/σ仅统计有效天(B列非空;参与天数=T)；'
             'LCB(悲观排名)=μ+z·σ/√T(z=B3=1)，越小越好(对齐前端879行)。')
    line8 = (f'8. 本表A-D列取自真实报文组合“{target["label"]}”({dates[0]}~{dates[-1]})，'
             f'C/D列=报文原值(单位:元,显示2位小数)；B列每日排名按{args.rank_mode}口径计算'
             f'(当日收益降序、并列同名次跳号的竞赛排名法)；报文基准validDays={target["validDays"]}'
             f'/winDays={target["winDays"]}/winRate={target["winRate"]:.4f}/coverage={target["coverage"]:.4f}')
    if ws[NOTE_TITLE_CELL].value == '口径说明':
        merged_note = any(str(m) == 'A52:E59' for m in ws.merged_cells.ranges)
        if merged_note:
            # 合并模板：全量文本写在 A52 锚点，按序号替换第3/4/6/8条(第6条负责校正LCB公式口径)
            import re
            text = ws['A52'].value or ''
            for line in (line3, line4, line6, line8):
                num = line[0]
                pat = re.compile(rf'{num}\. .*?(?=\n\d+\. |\Z)', re.S)
                text = pat.sub(line, text, count=1) if pat.search(text) else (text + '\n' + line if text else line)
            ws['A52'] = text
        else:
            ws['A54'], ws['A55'], ws['A59'] = line3, line4, line8
    try:
        wb.save(excel_path)
    except PermissionError:
        sys.exit(f'[ERROR] Excel 文件被占用，请先关闭后重试: {excel_path}')
    print(f'已写入 {excel_path}')
    print(f'高亮无效天行: {hl_rows if hl_rows else "无"} | A5={title_text} | B2 已同步为 {len(combos)} | '
          f'B4 已同步为 {n} | E列公式已重建 | B16={tag} | B17={quant_text}')

    # ---- 7. 落盘回读校验（重新打开磁盘文件；历史教训：写入被中断时可能只落盘部分列） ----
    wb2 = openpyxl.load_workbook(excel_path)
    ws2 = wb2.active
    bad = []
    for i, dt in enumerate(dates):
        r = DATA_START + i
        got = tuple(ws2.cell(row=r, column=j).value for j in (1, 2, 3, 4))
        exp = (dt, ranks.get(dt), sp_of(target, dt), base.get(dt))
        # C/D列浮点比对用容差：Excel存取会做二进制浮点往返舍入(末位1e-12级差异属正常)
        equal = got[:2] == exp[:2] and all(
            g is None and e is None or
            isinstance(g, (int, float)) and isinstance(e, (int, float)) and abs(g - e) < 1e-9
            for g, e in zip(got[2:], exp[2:]))
        if not equal:
            bad.append((dt, got, exp))
    # A/B/C/D 全量精确比对已覆盖所有列（含无效天应为空的B/C/D），无需额外非空规则
    e_ok = all(str(ws2.cell(row=DATA_START + i, column=5).value or '').startswith('=IF(') for i in range(n))
    sim_t = sum(1 for i, dt in enumerate(dates)
                if all(ws2.cell(row=DATA_START + i, column=j).value not in (None, '')
                       for j in (2, 4))
                and ws2.cell(row=DATA_START + i, column=3).value not in (None, '', 0))
    status = 'PASS' if (not bad and e_ok and sim_t == t and ws2[TAG_CELL].value == tag
                        and ws2[QUANT_CELL].value == quant_text
                        and ws2[TITLE_CELL].value == title_text) else 'FAIL'
    print(f'落盘校验: A/B/C/D {n}行全量比对 {"PASS" if not bad else bad[:3]} | E列公式 {"PASS" if e_ok else "FAIL"} | '
          f'A5标题 {"PASS" if ws2[TITLE_CELL].value == title_text else "FAIL"} | '
          f'B16标签 {"PASS" if ws2[TAG_CELL].value == tag else "FAIL"} | '
          f'B17分位线 {"PASS" if ws2[QUANT_CELL].value == quant_text else "FAIL"} | '
          f'模拟E列T={sim_t} (应{t}) {"PASS" if sim_t == t else "FAIL"} => {status}')
    if status == 'FAIL':
        print('[ERROR] 落盘校验失败，请确认 Excel 已关闭且未被其他程序修改后重跑', file=sys.stderr)
        sys.exit(1)
    print('全部完成 ✔')


if __name__ == '__main__':
    main()
